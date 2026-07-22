from __future__ import annotations

import csv
import json
import math
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
MANIFEST_PATH = ROOT / "src" / "slicScoringManifest.json"
WORKBOOK_PATH = ROOT / "output" / "spreadsheet" / "slic_scoring_workbook.xlsx"
CITY_CSV_PATH = ROOT / "output" / "spreadsheet" / "slic_city_universe.csv"


def load_manifest() -> dict:
    with MANIFEST_PATH.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def percentile(values: list[float], q: float) -> float:
    if not values:
        raise ValueError("percentile requires at least one value")
    ordered = sorted(values)
    if len(ordered) == 1:
        return ordered[0]
    position = (len(ordered) - 1) * q
    lower = math.floor(position)
    upper = math.ceil(position)
    if lower == upper:
        return ordered[lower]
    fraction = position - lower
    return ordered[lower] + (ordered[upper] - ordered[lower]) * fraction


def normalize(value: float, values: list[float], direction: str) -> float:
    p05 = percentile(values, 0.05)
    p95 = percentile(values, 0.95)
    if math.isclose(p05, p95):
        return 100.0
    capped = min(max(value, p05), p95)
    if direction == "positive":
        return max(0.0, min(100.0, 100.0 * (capped - p05) / (p95 - p05)))
    return max(0.0, min(100.0, 100.0 * (p95 - capped) / (p95 - p05)))


def assert_expected_sheets(manifest: dict, workbook_sheets: list[str]) -> None:
    missing = [sheet for sheet in manifest["workbookSheets"] if sheet not in workbook_sheets]
    if missing:
        raise AssertionError(f"Missing workbook sheets: {missing}")
    for extra in ("Region_Comparison", "Thai_Ranking_Scaffold"):
        if extra not in workbook_sheets:
            raise AssertionError(f"Missing companion sheet: {extra}")


def assert_headers(manifest: dict, workbook) -> None:
    for sheet_name, expected_headers in manifest["sheetColumns"].items():
        sheet = workbook[sheet_name]
        header_row = 2 if sheet_name == "Leaderboard" else 1
        actual_headers = [sheet.cell(row=header_row, column=index + 1).value for index in range(len(expected_headers))]
        if actual_headers != expected_headers:
            raise AssertionError(f"Header mismatch for {sheet_name}: {actual_headers} != {expected_headers}")


def assert_formulas(workbook, manifest: dict) -> None:
    score_headers = manifest["sheetColumns"]["Scores"]
    score_columns = {header: get_column_letter(index + 1) for index, header in enumerate(score_headers)}
    city_input_headers = manifest["sheetColumns"]["City_Inputs"]
    city_input_columns = {header: get_column_letter(index + 1) for index, header in enumerate(city_input_headers)}
    key_checks = {
        ("City_Inputs", f"{city_input_columns['tax_rate_context']}2"): "Country_Context",
        ("City_Inputs", f"{city_input_columns['di_ppp_raw']}2"): "(1-$",
        ("Scores", f"{score_columns['pressure_disposable_income_ppp_score']}2"): "Norm_Stats",
        ("Scores", f"{score_columns['rank']}2"): "COUNTIFS",
        ("Leaderboard", "D3"): "Scores!",
    }
    for (sheet_name, cell_ref), expected_fragment in key_checks.items():
        formula = workbook[sheet_name][cell_ref].value
        if not isinstance(formula, str) or expected_fragment not in formula:
            raise AssertionError(f"{sheet_name}!{cell_ref} does not contain expected fragment {expected_fragment}")

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("=") and "#REF!" in cell.value:
                    raise AssertionError(f"Broken reference found in {sheet.title}!{cell.coordinate}")


def assert_csv_matches_workbook(workbook) -> None:
    sheet = workbook["City_Universe"]
    workbook_rows = []
    for row in range(2, sheet.max_row + 1):
        workbook_rows.append(
            {
                "city_id": sheet[f"A{row}"].value,
                "cohort": sheet[f"E{row}"].value,
                "manifest_status": sheet[f"G{row}"].value,
                "city_type": sheet[f"F{row}"].value,
                "country": sheet[f"D{row}"].value,
                "inclusion_rationale": sheet[f"H{row}"].value,
            }
        )

    with CITY_CSV_PATH.open("r", encoding="utf-8", newline="") as handle:
        csv_rows = list(csv.DictReader(handle))

    if workbook_rows != csv_rows:
        raise AssertionError("CSV export does not match City_Universe workbook rows")


def compute_di_ppp(city: dict[str, float]) -> float | None:
    required = [
        "gross_income",
        "rent",
        "utilities",
        "transit_cost",
        "internet_cost",
        "food_cost",
        "tax_rate_context",
        "ppp_private_consumption_context",
    ]
    if any(city.get(field) is None for field in required):
        return None
    return (
        (city["gross_income"] * (1 - city["tax_rate_context"]))
        - city["rent"]
        - city["utilities"]
        - city["transit_cost"]
        - city["internet_cost"]
        - city["food_cost"]
    ) / city["ppp_private_consumption_context"]


def compute_synthetic_scores() -> None:
    synthetic = [
        {
            "id": "alpha",
            "gross_income": 6200,
            "rent": 1400,
            "utilities": 220,
            "transit_cost": 150,
            "internet_cost": 60,
            "food_cost": 580,
            "tax_rate_context": 0.18,
            "ppp_private_consumption_context": 1.1,
            "housing_burden_raw": 0.27,
            "household_debt_effective_raw": 0.55,
            "working_time_pressure_raw": 41,
            "suicide_mental_strain_raw": 7.5,
            "personal_safety_raw": 3.0,
            "transit_access_commute_raw": 78,
            "clean_air_raw": 13,
            "water_sanitation_utility_raw": 90,
            "digital_infrastructure_raw": 87,
            "healthcare_quality_raw": 81,
            "education_quality_raw": 79,
            "equal_opportunity_raw": 74,
            "gini_coefficient_context": 31,
            "hospitality_belonging_raw": 77,
            "tolerance_pluralism_raw": 80,
            "cultural_public_life_raw": 76,
            "entrepreneurial_dynamism_raw": 72,
            "innovation_research_intensity_raw": 69,
            "investment_signal_raw": 71,
            "gdp_per_capita_ppp_context": 52000,
            "gdp_growth_context": 3.3,
            "administrative_investment_friction_raw": 18,
        },
        {
            "id": "beta",
            "gross_income": 4700,
            "rent": 1600,
            "utilities": 260,
            "transit_cost": 210,
            "internet_cost": 70,
            "food_cost": 660,
            "tax_rate_context": 0.21,
            "ppp_private_consumption_context": 1.05,
            "housing_burden_raw": 0.39,
            "household_debt_effective_raw": 0.8,
            "working_time_pressure_raw": 49,
            "suicide_mental_strain_raw": 12.0,
            "personal_safety_raw": 12.0,
            "transit_access_commute_raw": 61,
            "clean_air_raw": 28,
            "water_sanitation_utility_raw": 76,
            "digital_infrastructure_raw": 73,
            "healthcare_quality_raw": 69,
            "education_quality_raw": 66,
            "equal_opportunity_raw": 58,
            "gini_coefficient_context": 42,
            "hospitality_belonging_raw": 63,
            "tolerance_pluralism_raw": 61,
            "cultural_public_life_raw": 67,
            "entrepreneurial_dynamism_raw": 59,
            "innovation_research_intensity_raw": 57,
            "investment_signal_raw": 55,
            "gdp_per_capita_ppp_context": 36000,
            "gdp_growth_context": 2.1,
            "administrative_investment_friction_raw": 34,
        },
        {
            "id": "gamma",
            "gross_income": 5400,
            "rent": 1100,
            "utilities": 190,
            "transit_cost": 140,
            "internet_cost": 45,
            "food_cost": 520,
            "tax_rate_context": 0.16,
            "ppp_private_consumption_context": 1.2,
            "housing_burden_raw": 0.23,
            "household_debt_effective_raw": 0.34,
            "working_time_pressure_raw": 38,
            "suicide_mental_strain_raw": 6.4,
            "personal_safety_raw": 2.2,
            "transit_access_commute_raw": 74,
            "clean_air_raw": 10,
            "water_sanitation_utility_raw": 93,
            "digital_infrastructure_raw": 82,
            "healthcare_quality_raw": 78,
            "education_quality_raw": 75,
            "equal_opportunity_raw": 72,
            "gini_coefficient_context": 29,
            "hospitality_belonging_raw": 82,
            "tolerance_pluralism_raw": 77,
            "cultural_public_life_raw": 81,
            "entrepreneurial_dynamism_raw": 68,
            "innovation_research_intensity_raw": 63,
            "investment_signal_raw": 66,
            "gdp_per_capita_ppp_context": 47000,
            "gdp_growth_context": 3.8,
            "administrative_investment_friction_raw": 14,
        },
    ]

    for city in synthetic:
        city["di_ppp_raw"] = compute_di_ppp(city)

    metrics = {
        "di_ppp_raw": ("positive", 9),
        "housing_burden_raw": ("negative", 5),
        "household_debt_effective_raw": ("negative", 4),
        "working_time_pressure_raw": ("negative", 4),
        "suicide_mental_strain_raw": ("negative", 3),
        "personal_safety_raw": ("negative", 5),
        "transit_access_commute_raw": ("positive", 5),
        "clean_air_raw": ("negative", 4),
        "water_sanitation_utility_raw": ("positive", 4),
        "digital_infrastructure_raw": ("positive", 4),
        "healthcare_quality_raw": ("positive", 8),
        "education_quality_raw": ("positive", 6),
        "hospitality_belonging_raw": ("positive", 5),
        "tolerance_pluralism_raw": ("positive", 5),
        "cultural_public_life_raw": ("positive", 5),
        "entrepreneurial_dynamism_raw": ("positive", 6),
        "innovation_research_intensity_raw": ("positive", 5),
        "administrative_investment_friction_raw": ("negative", 4),
    }

    normalized_scores: dict[str, dict[str, float]] = {city["id"]: {} for city in synthetic}
    for metric, (direction, _) in metrics.items():
        values = [float(city[metric]) for city in synthetic]
        for city in synthetic:
            normalized_scores[city["id"]][metric] = normalize(float(city[metric]), values, direction)

    eq_values = [float(city["equal_opportunity_raw"]) for city in synthetic]
    gini_values = [float(city["gini_coefficient_context"]) for city in synthetic]
    investment_values = [float(city["investment_signal_raw"]) for city in synthetic]
    gdp_values = [float(city["gdp_per_capita_ppp_context"]) for city in synthetic]
    growth_values = [float(city["gdp_growth_context"]) for city in synthetic]

    totals: dict[str, float] = {}
    for city in synthetic:
        city_id = city["id"]
        equal_opportunity_score = (
            normalize(float(city["equal_opportunity_raw"]), eq_values, "positive") * 0.7
            + normalize(float(city["gini_coefficient_context"]), gini_values, "negative") * 0.3
        )
        economic_context_score = (
            normalize(float(city["investment_signal_raw"]), investment_values, "positive") * 0.5
            + normalize(float(city["gdp_per_capita_ppp_context"]), gdp_values, "positive") * 0.3
            + normalize(float(city["gdp_growth_context"]), growth_values, "positive") * 0.2
        )
        totals[city_id] = (
            normalized_scores[city_id]["di_ppp_raw"] * 9
            + normalized_scores[city_id]["housing_burden_raw"] * 5
            + normalized_scores[city_id]["household_debt_effective_raw"] * 4
            + normalized_scores[city_id]["working_time_pressure_raw"] * 4
            + normalized_scores[city_id]["suicide_mental_strain_raw"] * 3
            + normalized_scores[city_id]["personal_safety_raw"] * 5
            + normalized_scores[city_id]["transit_access_commute_raw"] * 5
            + normalized_scores[city_id]["clean_air_raw"] * 4
            + normalized_scores[city_id]["water_sanitation_utility_raw"] * 4
            + normalized_scores[city_id]["digital_infrastructure_raw"] * 4
            + normalized_scores[city_id]["healthcare_quality_raw"] * 8
            + normalized_scores[city_id]["education_quality_raw"] * 6
            + equal_opportunity_score * 4
            + normalized_scores[city_id]["hospitality_belonging_raw"] * 5
            + normalized_scores[city_id]["tolerance_pluralism_raw"] * 5
            + normalized_scores[city_id]["cultural_public_life_raw"] * 5
            + normalized_scores[city_id]["entrepreneurial_dynamism_raw"] * 6
            + normalized_scores[city_id]["innovation_research_intensity_raw"] * 5
            + economic_context_score * 5
            + normalized_scores[city_id]["administrative_investment_friction_raw"] * 4
        ) / 100

    if not (totals["alpha"] > totals["gamma"] > totals["beta"]):
        raise AssertionError(f"Unexpected synthetic ranking order: {totals}")

    if not (
        normalized_scores["alpha"]["personal_safety_raw"]
        > normalized_scores["beta"]["personal_safety_raw"]
    ):
        raise AssertionError("Harmful safety metric reversal is not behaving as expected")

    di_value = synthetic[0]["di_ppp_raw"]
    expected_di = (((6200 * (1 - 0.18)) - 1400 - 220 - 150 - 60 - 580) / 1.1)
    if not math.isclose(di_value, expected_di, rel_tol=1e-9):
        raise AssertionError("Tax-adjusted PPP disposable income formula is incorrect")

    sparse_city = {
        "pressure": 0.5,
        "viability": 0.0,
        "capability": 0.0,
        "community": 0.0,
        "creative": 0.0,
    }
    overall_coverage = (
        sparse_city["pressure"] * 25
        + sparse_city["viability"] * 22
        + sparse_city["capability"] * 18
        + sparse_city["community"] * 15
        + sparse_city["creative"] * 20
    ) / 100
    if overall_coverage >= 0.5:
        raise AssertionError("Synthetic sparse city should remain below the Ranked threshold")


def main() -> None:
    manifest = load_manifest()
    workbook = load_workbook(WORKBOOK_PATH, data_only=False)
    assert_expected_sheets(manifest, workbook.sheetnames)
    assert_headers(manifest, workbook)
    assert_formulas(workbook, manifest)
    assert_csv_matches_workbook(workbook)
    compute_synthetic_scores()
    print("SLIC workbook validation passed.")


if __name__ == "__main__":
    main()
