from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


ROOT = Path(__file__).resolve().parents[1]
MANIFEST_PATH = ROOT / "src" / "slicScoringManifest.json"
OUTPUT_DIR = ROOT / "output" / "spreadsheet"
WORKBOOK_PATH = OUTPUT_DIR / "slic_scoring_workbook.xlsx"
CITY_CSV_PATH = OUTPUT_DIR / "slic_city_universe.csv"
APP_CITY_CSV_PATH = ROOT / "src" / "data" / "slic_city_universe.csv"
REGION_COMPARISON_SHEET = "Region_Comparison"
THAI_SCAFFOLD_SHEET = "Thai_Ranking_Scaffold"


HEADER_FILL = PatternFill("solid", fgColor="D9E2EC")
SUBHEADER_FILL = PatternFill("solid", fgColor="F5F7FA")
INPUT_FILL = PatternFill("solid", fgColor="E8F1FF")
FORMULA_FILL = PatternFill("solid", fgColor="F3F0FF")
CONTEXT_FILL = PatternFill("solid", fgColor="E8F7F1")
WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
TITLE_FILL = PatternFill("solid", fgColor="EEF2F5")
THIN_BORDER = Border(bottom=Side(style="thin", color="CBD2D9"))
HEADER_FONT = Font(bold=True, color="1F2933")
TITLE_FONT = Font(bold=True, size=14, color="1F2933")
WRAP = Alignment(wrap_text=True, vertical="top")


def load_manifest() -> dict:
    with MANIFEST_PATH.open("r", encoding="utf-8") as handle:
        return json.load(handle)


MANIFEST = load_manifest()


def row(
    city_id: str,
    display_name: str,
    fua_name: str,
    country: str,
    cohort: str,
    city_type: str,
    manifest_status: str,
    inclusion_rationale: str,
) -> dict[str, str]:
    return {
        "city_id": city_id,
        "display_name": display_name,
        "fua_name": fua_name,
        "country": country,
        "cohort": cohort,
        "city_type": city_type,
        "manifest_status": manifest_status,
        "inclusion_rationale": inclusion_rationale,
    }


def build_city_universe() -> list[dict[str, str]]:
    return [
        row(
            "th-bangkok",
            "Bangkok",
            "Bangkok Metropolitan Region",
            "Thailand",
            "Southeast Asia",
            "primary",
            "locked",
            "User anchor and regional benchmark for hospitality, economic dynamism, and affordability-pressure tradeoffs.",
        ),
        row(
            "sg-singapore",
            "Singapore",
            "Singapore",
            "Singapore",
            "Southeast Asia",
            "primary",
            "locked",
            "Global convenience benchmark included to test how systems quality interacts with pressure, cost, and meaning.",
        ),
        row(
            "my-kuala-lumpur",
            "Kuala Lumpur",
            "Klang Valley",
            "Malaysia",
            "Southeast Asia",
            "primary",
            "locked",
            "User anchor and regional economic core with strong competitiveness, connectivity, and mixed affordability.",
        ),
        row(
            "id-jakarta",
            "Jakarta",
            "Jakarta Metropolitan Area",
            "Indonesia",
            "Southeast Asia",
            "primary",
            "provisional",
            "Necessary Southeast Asian benchmark for scale, competitiveness, and infrastructure-pressure tradeoffs.",
        ),
        row(
            "ph-makati",
            "Makati",
            "Metro Manila Core Business District",
            "Philippines",
            "Southeast Asia",
            "primary",
            "provisional",
            "Metro-core Philippine test case included for business density, services, and lived-affordability stress.",
        ),
        row(
            "my-george-town",
            "George Town",
            "Penang Conurbation",
            "Malaysia",
            "Southeast Asia",
            "secondary",
            "locked",
            "User anchor for semiconductor economy, cultural texture, and secondary-city livability.",
        ),
        row(
            "th-phuket",
            "Phuket",
            "Phuket Island Urban Area",
            "Thailand",
            "Southeast Asia",
            "secondary",
            "locked",
            "User anchor and tourism-livability stress test for cost, safety, and service quality.",
        ),
        row(
            "th-chiang-mai",
            "Chiang Mai",
            "Chiang Mai Metropolitan Area",
            "Thailand",
            "Southeast Asia",
            "secondary",
            "locked",
            "User anchor for creativity, secondary-city ambition, and lifestyle viability.",
        ),
        row(
            "my-kuching",
            "Kuching",
            "Greater Kuching",
            "Malaysia",
            "Southeast Asia",
            "secondary",
            "locked",
            "User anchor included as a calmer secondary-city comparator with strong everyday livability appeal.",
        ),
        row(
            "th-hat-yai",
            "Hat Yai",
            "Hat Yai-Songkhla Urban Area",
            "Thailand",
            "Southeast Asia",
            "secondary",
            "provisional",
            "Southern Thai trade and services challenger included to test regional depth beyond capital-city bias.",
        ),
        row(
            "tw-taipei",
            "Taipei",
            "Taipei-Keelung-Taoyuan Metropolitan Area",
            "Taiwan",
            "East Asia",
            "primary",
            "locked",
            "High-systems, high-street-life benchmark aligned with earlier SLIC discussion.",
        ),
        row(
            "kr-busan",
            "Busan",
            "Greater Busan",
            "South Korea",
            "East Asia",
            "primary",
            "locked",
            "User anchor for second-city infrastructure quality, coastal livability, and economic competence.",
        ),
        row(
            "cn-shanghai",
            "Shanghai",
            "Shanghai Metropolitan Area",
            "China",
            "East Asia",
            "primary",
            "provisional",
            "Economic gravity benchmark included to test how mega-city capability meets affordability and civic-pressure tradeoffs.",
        ),
        row(
            "cn-shenzhen",
            "Shenzhen",
            "Shenzhen Metropolitan Area",
            "China",
            "East Asia",
            "primary",
            "provisional",
            "Innovation and manufacturing benchmark included for productive economic strength and high-pressure urban modernity.",
        ),
        row(
            "jp-hiroshima",
            "Hiroshima",
            "Hiroshima Metropolitan Employment Area",
            "Japan",
            "East Asia",
            "primary",
            "provisional",
            "Provisional Japanese primary candidate with stronger everyday scale than megacity peers.",
        ),
        row(
            "kr-jeju-city",
            "Jeju City",
            "Jeju Urban Area",
            "South Korea",
            "East Asia",
            "secondary",
            "locked",
            "User anchor for lifestyle quality, hospitality, and island-scale livability.",
        ),
        row(
            "jp-fukuoka",
            "Fukuoka",
            "Fukuoka Metropolitan Area",
            "Japan",
            "East Asia",
            "secondary",
            "locked",
            "User anchor and strong challenger city for innovation, access, and livability balance.",
        ),
        row(
            "jp-kobe",
            "Kobe",
            "Kobe Metropolitan Area",
            "Japan",
            "East Asia",
            "secondary",
            "locked",
            "User anchor included as a port-city benchmark with strong urban form and quality-of-life reputation.",
        ),
        row(
            "jp-sapporo",
            "Sapporo",
            "Sapporo Metropolitan Area",
            "Japan",
            "East Asia",
            "secondary",
            "locked",
            "User anchor and climate-diversified challenger with strong everyday manageability.",
        ),
        row(
            "cn-tianjin",
            "Tianjin",
            "Tianjin Metropolitan Area",
            "China",
            "East Asia",
            "secondary",
            "provisional",
            "Secondary East Asian challenger included for industrial relevance, port-city utility, and relief from capital-city prestige bias.",
        ),
        row(
            "in-bengaluru",
            "Bengaluru",
            "Bengaluru Metropolitan Area",
            "India",
            "South Asia",
            "primary",
            "locked",
            "Primary South Asian innovation benchmark included despite affordability and infrastructure pressures.",
        ),
        row(
            "in-hyderabad",
            "Hyderabad",
            "Hyderabad Metropolitan Region",
            "India",
            "South Asia",
            "primary",
            "locked",
            "Paired with Bengaluru as a competitiveness and opportunity benchmark for South Asia.",
        ),
        row(
            "lk-colombo",
            "Colombo",
            "Colombo Metropolitan Region",
            "Sri Lanka",
            "South Asia",
            "primary",
            "locked",
            "Regional capital-city comparator with manageable scale and stronger coastal livability potential.",
        ),
        row(
            "bd-dhaka",
            "Dhaka",
            "Dhaka Metropolitan Area",
            "Bangladesh",
            "South Asia",
            "primary",
            "provisional",
            "High-pressure megacity included provisionally as a contrast case inside the South Asian universe.",
        ),
        row(
            "pk-karachi",
            "Karachi",
            "Karachi Metropolitan Area",
            "Pakistan",
            "South Asia",
            "primary",
            "provisional",
            "Provisional large-metro comparator included to keep the South Asian candidate field honest and broad.",
        ),
        row(
            "pk-lahore",
            "Lahore",
            "Lahore Metropolitan Area",
            "Pakistan",
            "South Asia",
            "secondary",
            "locked",
            "Secondary challenger included for cultural weight and regional importance.",
        ),
        row(
            "bd-chattogram",
            "Chattogram",
            "Chattogram Metropolitan Area",
            "Bangladesh",
            "South Asia",
            "secondary",
            "locked",
            "Secondary port-city comparator for industrial relevance and everyday-scale testing.",
        ),
        row(
            "np-kathmandu",
            "Kathmandu",
            "Kathmandu Valley",
            "Nepal",
            "South Asia",
            "secondary",
            "provisional",
            "Provisional cultural and tourism-weighted challenger with strong environmental tradeoffs.",
        ),
        row(
            "bt-thimphu",
            "Thimphu",
            "Thimphu Urban Area",
            "Bhutan",
            "South Asia",
            "secondary",
            "locked",
            "Smaller-scale South Asian challenger included for governance, environment, and lifestyle contrast.",
        ),
        row(
            "mv-male",
            "Male",
            "Male Urban Region",
            "Maldives",
            "South Asia",
            "secondary",
            "locked",
            "Island-city comparator included for service quality, compactness, and extreme spatial constraints.",
        ),
        row(
            "fr-paris",
            "Paris",
            "Paris Metropolitan Area",
            "France",
            "Western and Northern Europe",
            "primary",
            "locked",
            "User anchor and global cultural capital included as a prestige city that SLIC can re-score through affordability and pressure.",
        ),
        row(
            "at-vienna",
            "Vienna",
            "Vienna Metropolitan Area",
            "Austria",
            "Western and Northern Europe",
            "primary",
            "locked",
            "User anchor included as a conventional benchmark that SLIC can test against pressure and creative vitality.",
        ),
        row(
            "ch-zurich",
            "Zurich",
            "Zurich Metropolitan Area",
            "Switzerland",
            "Western and Northern Europe",
            "primary",
            "locked",
            "High-convenience benchmark included so SLIC can visibly penalize cost where necessary.",
        ),
        row(
            "nl-amsterdam",
            "Amsterdam",
            "Amsterdam Metropolitan Area",
            "Netherlands",
            "Western and Northern Europe",
            "primary",
            "provisional",
            "Provisional benchmark for global talent pull, openness, and cost pressure.",
        ),
        row(
            "dk-copenhagen",
            "Copenhagen",
            "Copenhagen Metropolitan Area",
            "Denmark",
            "Western and Northern Europe",
            "primary",
            "provisional",
            "Provisional systems-quality comparator included for mobility, governance, and cost tradeoffs.",
        ),
        row(
            "at-graz",
            "Graz",
            "Graz Urban Region",
            "Austria",
            "Western and Northern Europe",
            "secondary",
            "locked",
            "User anchor and secondary Austrian challenger with lower prestige bias than Vienna.",
        ),
        row(
            "pt-porto",
            "Porto",
            "Porto Metropolitan Area",
            "Portugal",
            "Western and Northern Europe",
            "secondary",
            "locked",
            "Secondary-city benchmark for cultural pull, walkability, and relative affordability.",
        ),
        row(
            "pt-braga",
            "Braga",
            "Braga Urban Area",
            "Portugal",
            "Western and Northern Europe",
            "secondary",
            "locked",
            "Secondary Portuguese challenger with strong everyday livability and university-city energy.",
        ),
        row(
            "ee-tallinn",
            "Tallinn",
            "Tallinn Urban Area",
            "Estonia",
            "Western and Northern Europe",
            "secondary",
            "locked",
            "Digitally competent northern-European challenger with manageable scale.",
        ),
        row(
            "fi-helsinki",
            "Helsinki",
            "Helsinki Metropolitan Area",
            "Finland",
            "Western and Northern Europe",
            "secondary",
            "provisional",
            "Provisional Nordic comparator included for systems quality, climate, and cost balance.",
        ),
        row(
            "ru-moscow",
            "Moscow",
            "Moscow Metropolitan Area",
            "Russia",
            "Southern/Eastern Europe and Eurasia",
            "primary",
            "locked",
            "User anchor included for scale, culture, infrastructure depth, and Eurasian comparability.",
        ),
        row(
            "hu-budapest",
            "Budapest",
            "Budapest Metropolitan Area",
            "Hungary",
            "Southern/Eastern Europe and Eurasia",
            "primary",
            "locked",
            "User anchor with strong urban form, culture, and affordability relative to western peers.",
        ),
        row(
            "pl-krakow",
            "Krakow",
            "Krakow Metropolitan Area",
            "Poland",
            "Southern/Eastern Europe and Eurasia",
            "primary",
            "locked",
            "User anchor and strong central-European benchmark for heritage, education, and secondary-scale quality.",
        ),
        row(
            "ro-bucharest",
            "Bucharest",
            "Bucharest Metropolitan Area",
            "Romania",
            "Southern/Eastern Europe and Eurasia",
            "primary",
            "locked",
            "Regional capital included for affordability-competitiveness balance and infrastructure trajectory.",
        ),
        row(
            "rs-belgrade",
            "Belgrade",
            "Belgrade Metropolitan Area",
            "Serbia",
            "Southern/Eastern Europe and Eurasia",
            "primary",
            "locked",
            "Regional challenger included for dynamism, hospitality, and lower-cost competitiveness.",
        ),
        row(
            "ru-nizhny-novgorod",
            "Nizhny Novgorod",
            "Nizhny Novgorod Urban Agglomeration",
            "Russia",
            "Southern/Eastern Europe and Eurasia",
            "secondary",
            "provisional",
            "User-referenced Russian challenger included provisionally as a second-tier Eurasian comparator.",
        ),
        row(
            "pl-katowice",
            "Katowice",
            "Upper Silesian Metropolitan Association",
            "Poland",
            "Southern/Eastern Europe and Eurasia",
            "secondary",
            "locked",
            "User anchor included for industrial-transition competitiveness and metro-scale livability.",
        ),
        row(
            "pl-gdansk",
            "Gdansk",
            "Tricity Metropolitan Area",
            "Poland",
            "Southern/Eastern Europe and Eurasia",
            "secondary",
            "locked",
            "User anchor included for port-city vitality and Baltic urban quality.",
        ),
        row(
            "pl-torun",
            "Torun",
            "Torun Urban Area",
            "Poland",
            "Southern/Eastern Europe and Eurasia",
            "secondary",
            "provisional",
            "User-referenced university-city challenger retained provisionally for follow-up data testing.",
        ),
        row(
            "sk-bratislava",
            "Bratislava",
            "Bratislava Metropolitan Area",
            "Slovakia",
            "Southern/Eastern Europe and Eurasia",
            "secondary",
            "provisional",
            "Secondary central-European challenger included for affordability, cross-border access, and competitiveness.",
        ),
        row(
            "us-chicago",
            "Chicago",
            "Chicago Metropolitan Area",
            "United States",
            "North America",
            "primary",
            "locked",
            "U.S. control city included for scale, connectivity, and relative affordability within the American field.",
        ),
        row(
            "mx-mexico-city",
            "Mexico City",
            "Valley of Mexico Metropolitan Area",
            "Mexico",
            "North America",
            "primary",
            "locked",
            "North American megacity benchmark included for urban energy, culture, and affordability tension.",
        ),
        row(
            "ca-toronto",
            "Toronto",
            "Greater Toronto Area",
            "Canada",
            "North America",
            "primary",
            "provisional",
            "High-cost North American comparator included to test whether strong services can offset pressure.",
        ),
        row(
            "pa-panama-city",
            "Panama City",
            "Panama City Metropolitan Area",
            "Panama",
            "North America",
            "primary",
            "locked",
            "Regional gateway benchmark included for connectivity, investment, and logistics-driven vitality.",
        ),
        row(
            "cr-san-jose",
            "San Jose",
            "Greater Metropolitan Area of Costa Rica",
            "Costa Rica",
            "North America",
            "primary",
            "locked",
            "Smaller North American capital benchmark included for education, healthcare, and relative stability.",
        ),
        row(
            "us-pittsburgh",
            "Pittsburgh",
            "Pittsburgh Metropolitan Area",
            "United States",
            "North America",
            "secondary",
            "locked",
            "Secondary U.S. challenger included for education, healthcare, and value relative to coastal hubs.",
        ),
        row(
            "ca-montreal",
            "Montreal",
            "Montreal Metropolitan Area",
            "Canada",
            "North America",
            "secondary",
            "locked",
            "Secondary Canadian benchmark with strong culture, transit, and urban texture.",
        ),
        row(
            "mx-guadalajara",
            "Guadalajara",
            "Guadalajara Metropolitan Area",
            "Mexico",
            "North America",
            "secondary",
            "locked",
            "Secondary Mexican challenger for entrepreneurship, liveability, and cost competitiveness.",
        ),
        row(
            "pr-san-juan",
            "San Juan",
            "San Juan Metropolitan Area",
            "Puerto Rico",
            "North America",
            "secondary",
            "provisional",
            "Caribbean urban challenger included provisionally for climate, connectivity, and services testing.",
        ),
        row(
            "mx-merida",
            "Merida",
            "Merida Metropolitan Area",
            "Mexico",
            "North America",
            "secondary",
            "provisional",
            "Lower-pressure Mexican challenger included provisionally for safety and lifestyle contrast.",
        ),
        row(
            "br-sao-paulo",
            "Sao Paulo",
            "Sao Paulo Metropolitan Area",
            "Brazil",
            "Latin America",
            "primary",
            "locked",
            "User anchor and Latin American scale benchmark for competitiveness and opportunity.",
        ),
        row(
            "ar-buenos-aires",
            "Buenos Aires",
            "Greater Buenos Aires",
            "Argentina",
            "Latin America",
            "primary",
            "locked",
            "Primary benchmark for culture, human capital, and urban texture across Latin America.",
        ),
        row(
            "cl-santiago",
            "Santiago",
            "Greater Santiago",
            "Chile",
            "Latin America",
            "primary",
            "locked",
            "Regional systems-quality benchmark included to compare efficiency against cost and social strain.",
        ),
        row(
            "co-bogota",
            "Bogota",
            "Bogota Metropolitan Area",
            "Colombia",
            "Latin America",
            "primary",
            "locked",
            "Primary comparator for scale, transit reform, education, and competitiveness.",
        ),
        row(
            "pe-lima",
            "Lima",
            "Lima Metropolitan Area",
            "Peru",
            "Latin America",
            "primary",
            "provisional",
            "Provisional Pacific-capital comparator included for scale and coastal livability tradeoffs.",
        ),
        row(
            "br-curitiba",
            "Curitiba",
            "Curitiba Metropolitan Area",
            "Brazil",
            "Latin America",
            "secondary",
            "locked",
            "Secondary benchmark for planning reputation, mobility, and relative affordability.",
        ),
        row(
            "co-medellin",
            "Medellin",
            "Aburra Valley Metropolitan Area",
            "Colombia",
            "Latin America",
            "secondary",
            "locked",
            "Secondary challenger for entrepreneurship, hillside transit, and urban transformation.",
        ),
        row(
            "uy-montevideo",
            "Montevideo",
            "Montevideo Metropolitan Area",
            "Uruguay",
            "Latin America",
            "secondary",
            "locked",
            "Lower-pressure South American benchmark with strong civic and social indicators.",
        ),
        row(
            "ar-cordoba",
            "Cordoba",
            "Cordoba Metropolitan Area",
            "Argentina",
            "Latin America",
            "secondary",
            "provisional",
            "Secondary Argentine challenger included provisionally for education, cost, and regional depth.",
        ),
        row(
            "cl-valparaiso",
            "Valparaiso",
            "Greater Valparaiso",
            "Chile",
            "Latin America",
            "secondary",
            "provisional",
            "Port-city challenger included provisionally for culture and coastal public-life potential.",
        ),
        row(
            "ae-dubai",
            "Dubai",
            "Dubai Metropolitan Area",
            "United Arab Emirates",
            "Middle East",
            "primary",
            "locked",
            "Regional convenience and connectivity benchmark included to test cost and social texture tradeoffs.",
        ),
        row(
            "ae-abu-dhabi",
            "Abu Dhabi",
            "Abu Dhabi Metropolitan Area",
            "United Arab Emirates",
            "Middle East",
            "primary",
            "locked",
            "Systems-quality comparator with strong state capacity and high service competence.",
        ),
        row(
            "qa-doha",
            "Doha",
            "Doha Metropolitan Area",
            "Qatar",
            "Middle East",
            "primary",
            "locked",
            "High-capacity Gulf benchmark included for infrastructure, safety, and cost testing.",
        ),
        row(
            "sa-riyadh",
            "Riyadh",
            "Riyadh Metropolitan Area",
            "Saudi Arabia",
            "Middle East",
            "primary",
            "provisional",
            "Provisional primary comparator included for scale, investment, and administrative reform testing.",
        ),
        row(
            "il-tel-aviv",
            "Tel Aviv",
            "Tel Aviv Metropolitan Area",
            "Israel",
            "Middle East",
            "primary",
            "provisional",
            "Provisional innovation benchmark included for startup intensity, openness, and high cost.",
        ),
        row(
            "om-muscat",
            "Muscat",
            "Muscat Metropolitan Area",
            "Oman",
            "Middle East",
            "secondary",
            "locked",
            "Lower-pressure Gulf challenger with strong everyday service quality potential.",
        ),
        row(
            "bh-manama",
            "Manama",
            "Greater Manama",
            "Bahrain",
            "Middle East",
            "secondary",
            "locked",
            "Secondary Gulf comparator for openness, compactness, and service-led economy.",
        ),
        row(
            "sa-jeddah",
            "Jeddah",
            "Jeddah Metropolitan Area",
            "Saudi Arabia",
            "Middle East",
            "secondary",
            "locked",
            "Coastal Saudi challenger included for commerce, lifestyle, and second-city comparison.",
        ),
        row(
            "kw-kuwait-city",
            "Kuwait City",
            "Kuwait Metropolitan Area",
            "Kuwait",
            "Middle East",
            "secondary",
            "locked",
            "Compact Gulf benchmark included for service quality and economic strength.",
        ),
        row(
            "jo-amman",
            "Amman",
            "Greater Amman",
            "Jordan",
            "Middle East",
            "secondary",
            "provisional",
            "Provisional Levantine challenger included for education, services, and regional resilience.",
        ),
        row(
            "za-cape-town",
            "Cape Town",
            "Cape Town Metropolitan Area",
            "South Africa",
            "Africa",
            "primary",
            "locked",
            "African benchmark for environment, urban form, and inequality tradeoffs.",
        ),
        row(
            "za-johannesburg",
            "Johannesburg",
            "Johannesburg Metropolitan Area",
            "South Africa",
            "Africa",
            "primary",
            "provisional",
            "Provisional large-metro comparator included for competitiveness and safety-pressure contrasts.",
        ),
        row(
            "ke-nairobi",
            "Nairobi",
            "Nairobi Metropolitan Area",
            "Kenya",
            "Africa",
            "primary",
            "locked",
            "Primary East African benchmark for entrepreneurship, services, and regional gravity.",
        ),
        row(
            "rw-kigali",
            "Kigali",
            "Kigali Urban Area",
            "Rwanda",
            "Africa",
            "primary",
            "locked",
            "Primary challenger included for cleanliness, public order, and service reliability.",
        ),
        row(
            "ma-casablanca",
            "Casablanca",
            "Casablanca Metropolitan Area",
            "Morocco",
            "Africa",
            "primary",
            "provisional",
            "Provisional north-African benchmark for scale, port economy, and urban vitality.",
        ),
        row(
            "mu-port-louis",
            "Port Louis",
            "Port Louis Urban Area",
            "Mauritius",
            "Africa",
            "secondary",
            "locked",
            "Island-city challenger included for service quality, openness, and governance stability.",
        ),
        row(
            "bw-gaborone",
            "Gaborone",
            "Greater Gaborone",
            "Botswana",
            "Africa",
            "secondary",
            "locked",
            "Lower-pressure southern-African challenger with cleaner administrative profile.",
        ),
        row(
            "na-windhoek",
            "Windhoek",
            "Windhoek Urban Area",
            "Namibia",
            "Africa",
            "secondary",
            "locked",
            "Secondary challenger included for manageable scale and baseline service competence.",
        ),
        row(
            "gh-accra",
            "Accra",
            "Greater Accra Metropolitan Area",
            "Ghana",
            "Africa",
            "secondary",
            "locked",
            "West African benchmark for talent, services, and regional connectivity.",
        ),
        row(
            "sn-dakar",
            "Dakar",
            "Dakar Metropolitan Area",
            "Senegal",
            "Africa",
            "secondary",
            "provisional",
            "Provisional francophone west-African challenger with cultural and regional pull.",
        ),
        row(
            "nz-auckland",
            "Auckland",
            "Auckland Metropolitan Area",
            "New Zealand",
            "Oceania",
            "primary",
            "locked",
            "Primary Oceanian benchmark for human-capital quality, openness, and cost pressure.",
        ),
        row(
            "au-sydney",
            "Sydney",
            "Greater Sydney",
            "Australia",
            "Oceania",
            "primary",
            "provisional",
            "Global city control included so SLIC can test whether systems quality survives extreme cost pressure.",
        ),
        row(
            "au-melbourne",
            "Melbourne",
            "Greater Melbourne",
            "Australia",
            "Oceania",
            "primary",
            "provisional",
            "Prestige benchmark included to compare mainstream rankings against SLIC's pressure logic.",
        ),
        row(
            "au-brisbane",
            "Brisbane",
            "Greater Brisbane",
            "Australia",
            "Oceania",
            "primary",
            "locked",
            "Primary Australian challenger with lower cost pressure than Sydney and Melbourne.",
        ),
        row(
            "nz-wellington",
            "Wellington",
            "Wellington Metropolitan Area",
            "New Zealand",
            "Oceania",
            "primary",
            "locked",
            "Compact capital benchmark included for services, education, and public-life quality.",
        ),
        row(
            "au-perth",
            "Perth",
            "Greater Perth",
            "Australia",
            "Oceania",
            "secondary",
            "locked",
            "Western Australian challenger included for lifestyle, wages, and geographic isolation tradeoffs.",
        ),
        row(
            "au-adelaide",
            "Adelaide",
            "Greater Adelaide",
            "Australia",
            "Oceania",
            "secondary",
            "locked",
            "Secondary challenger with strong everyday scale and relatively lower housing pressure.",
        ),
        row(
            "au-hobart",
            "Hobart",
            "Greater Hobart",
            "Australia",
            "Oceania",
            "secondary",
            "locked",
            "Smaller-scale challenger included for lifestyle quality and public-environment strength.",
        ),
        row(
            "nz-christchurch",
            "Christchurch",
            "Christchurch Metropolitan Area",
            "New Zealand",
            "Oceania",
            "secondary",
            "provisional",
            "Provisional challenger included for rebuild-era infrastructure and lower-pressure urban living.",
        ),
        row(
            "fj-suva",
            "Suva",
            "Greater Suva",
            "Fiji",
            "Oceania",
            "secondary",
            "locked",
            "Pacific regional challenger included so Oceania is not reduced to Australia and New Zealand alone.",
        ),
    ]


THAI_SCAFFOLD_ROWS = [
    ("th-bangkok", "Bangkok", "Bangkok"),
    ("th-chonburi", "Chonburi-Pattaya", "Chonburi"),
    ("th-phuket", "Phuket", "Phuket"),
    ("th-chiang-mai", "Chiang Mai", "Chiang Mai"),
    ("th-chiang-rai", "Chiang Rai", "Chiang Rai"),
    ("th-nakhon-si-thammarat", "Nakhon Si Thammarat", "Nakhon Si Thammarat"),
    ("th-nakhon-ratchasima", "Nakhon Ratchasima", "Nakhon Ratchasima"),
    ("th-khon-kaen", "Khon Kaen", "Khon Kaen"),
    ("th-ubon-ratchathani", "Ubon Ratchathani", "Ubon Ratchathani"),
    ("th-surat-thani", "Surat Thani", "Surat Thani"),
    ("th-hat-yai", "Hat Yai", "Songkhla"),
]


def style_header_row(sheet: Worksheet, row_number: int) -> None:
    for cell in sheet[row_number]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER


def fit_columns(sheet: Worksheet, max_width: int = 28) -> None:
    widths: dict[int, int] = {}
    for row_cells in sheet.iter_rows():
        for cell in row_cells:
            if cell.value is None:
                continue
            value = str(cell.value)
            widths[cell.column] = min(max(widths.get(cell.column, 0), len(value) + 2), max_width)
    for column_index, width in widths.items():
        sheet.column_dimensions[get_column_letter(column_index)].width = max(width, 12)


def freeze_and_filter(sheet: Worksheet, freeze_at: str = "A2") -> None:
    sheet.freeze_panes = freeze_at
    if sheet.max_row >= 1 and sheet.max_column >= 1:
        sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"


def write_readme(
    sheet: Worksheet,
    manifest: dict,
) -> None:
    sheet["A1"] = "SLIC Scoring Workbook"
    sheet["A1"].font = TITLE_FONT
    sheet["A1"].fill = TITLE_FILL
    sheet["A1"].alignment = WRAP
    sheet.merge_cells("A1:F1")

    rows = [
        ("Purpose", "This workbook is the operational source of truth for the global SLIC 100-city ranking. Enter or import data into Country_Context and City_Inputs, then review Scores, Leaderboard, Region_Comparison, and the Thai-specific scaffold."),
        ("Official formula", "SLIC(c) = 0.25 Pressure + 0.22 DailyViability + 0.18 HumanCapability + 0.15 Community + 0.20 CreativeVitality"),
        ("Disposable income term", "DI_ppp = ((gross_income x (1 - tax_rate)) - rent - utilities - transit - internet - food) / ppp_private_consumption"),
        ("Normalization", "Each scored input is winsorized at the 5th and 95th percentiles, then converted to a 0-100 score. Harmful variables such as crime, debt, suicide, PM2.5, and administrative friction are reverse-scored."),
        ("Coverage rules", "A city is Ranked only when overall weighted coverage is at least 50% and every pillar clears the 35% minimum. Coverage grades: A >= 75%, B = 50-74%, C = 35-49%, below 35% = Watchlist."),
        ("Source hierarchy", "Tier 1 city/metro official data > Tier 2 subnational official data > Tier 3 national and international official data > Tier 4 audited secondary and experimental layers."),
        ("Doctrinal rule: safety", "Safety is scored by harm and victimization outcomes, never by camera count or visible control hardware."),
        ("Doctrinal rule: tolerance", "Tolerance is scored through low-friction coexistence and real personal freedom, not symbolic branding proxies."),
        ("Doctrinal rule: visitor flows", "Visitor flow is contextual only. It should be reviewed against crime, civic strain, and public-life quality rather than treated as an automatic positive."),
        ("Country context", "GDP per capita PPP, GDP growth, Gini coefficient, PPP private consumption, and user-supplied tax rate are visible in Country_Context. GDP terms are capped context inputs, not headline determinants."),
        ("How tax works", "Populate tax_rate_assumption in Country_Context as a decimal such as 0.18. City_Inputs pulls that value automatically into tax_rate_context and recomputes di_ppp_raw."),
        ("Region comparison", "Region_Comparison summarizes ranked city counts and average scores by cohort so Europe, Asia, the Americas, Africa, the Middle East, and Oceania can be compared cleanly."),
        ("Thai companion", "Thai_Ranking_Scaffold is a separate sheet for a Thailand-only ranking. It emphasizes province GDP per capita, pollution, hospitals, schools, infrastructure, income, longevity, and competitiveness."),
    ]

    start_row = 3
    for offset, (label, text) in enumerate(rows):
        row_num = start_row + offset
        sheet[f"A{row_num}"] = label
        sheet[f"A{row_num}"].font = HEADER_FONT
        sheet[f"A{row_num}"].fill = SUBHEADER_FILL
        sheet[f"A{row_num}"].alignment = WRAP
        sheet[f"B{row_num}"] = text
        sheet[f"B{row_num}"].alignment = WRAP
        sheet.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=6)

    threshold_row = start_row + len(rows) + 2
    sheet[f"A{threshold_row}"] = "Coverage thresholds"
    sheet[f"A{threshold_row}"].font = HEADER_FONT
    sheet[f"A{threshold_row}"].fill = SUBHEADER_FILL
    for idx, (key, value) in enumerate(manifest["coverageThresholds"].items(), start=threshold_row + 1):
        sheet[f"A{idx}"] = key
        sheet[f"B{idx}"] = value
        sheet[f"B{idx}"].number_format = "0%"

    tier_row = threshold_row + 8
    sheet[f"A{tier_row}"] = "Source tiers"
    sheet[f"A{tier_row}"].font = HEADER_FONT
    sheet[f"A{tier_row}"].fill = SUBHEADER_FILL
    for idx, tier in enumerate(manifest["sourceTiers"], start=tier_row + 1):
        sheet[f"A{idx}"] = tier["tier"]
        sheet[f"B{idx}"] = tier["title"]
        sheet[f"C{idx}"] = tier["description"]
        sheet[f"C{idx}"].alignment = WRAP

    for column in ("A", "B", "C", "D", "E", "F"):
        sheet.column_dimensions[column].width = 24 if column == "A" else 22


def write_table(sheet: Worksheet, headers: list[str], rows: list[dict[str, object]]) -> None:
    sheet.append(headers)
    style_header_row(sheet, 1)
    for record in rows:
        sheet.append([record.get(header, "") for header in headers])
    freeze_and_filter(sheet)
    fit_columns(sheet, max_width=36)


def unique_countries(city_universe: list[dict[str, str]]) -> list[str]:
    seen: list[str] = []
    for city in city_universe:
        country = city["country"]
        if country not in seen:
            seen.append(country)
    return seen


def header_map(headers: list[str]) -> dict[str, int]:
    return {header: index + 1 for index, header in enumerate(headers)}


def column_letter(headers: list[str], field: str) -> str:
    return get_column_letter(header_map(headers)[field])


def city_input_formula(
    row_number: int,
    country_context_column: str,
    country_context_rows: int,
    country_letter: str,
) -> str:
    return (
        f'=IFERROR(INDEX(Country_Context!${country_context_column}$2:${country_context_column}${country_context_rows},'
        f'MATCH(${country_letter}{row_number},Country_Context!$A$2:$A${country_context_rows},0)),"")'
    )


def build_city_inputs_sheet(
    sheet: Worksheet,
    headers: list[str],
    city_universe: list[dict[str, str]],
    country_headers: list[str],
    country_count: int,
) -> None:
    sheet.append(headers)
    style_header_row(sheet, 1)

    city_country_letter = column_letter(headers, "country")
    city_gross_income_letter = column_letter(headers, "gross_income")
    city_rent_letter = column_letter(headers, "rent")
    city_utilities_letter = column_letter(headers, "utilities")
    city_transit_cost_letter = column_letter(headers, "transit_cost")
    city_internet_cost_letter = column_letter(headers, "internet_cost")
    city_food_cost_letter = column_letter(headers, "food_cost")
    city_di_ppp_letter = column_letter(headers, "di_ppp_raw")
    city_household_debt_burden_letter = column_letter(headers, "household_debt_burden_raw")
    city_household_debt_effective_letter = column_letter(headers, "household_debt_effective_raw")
    city_gdp_context_letter = column_letter(headers, "gdp_per_capita_ppp_context")
    city_growth_context_letter = column_letter(headers, "gdp_growth_context")
    city_gini_context_letter = column_letter(headers, "gini_coefficient_context")
    city_ppp_context_letter = column_letter(headers, "ppp_private_consumption_context")
    city_tax_context_letter = column_letter(headers, "tax_rate_context")
    city_debt_context_letter = column_letter(headers, "household_debt_proxy_context")

    country_column_letters = {
        "gdp_per_capita_ppp_context": column_letter(country_headers, "gdp_per_capita_ppp"),
        "gdp_growth_context": column_letter(country_headers, "gdp_growth"),
        "gini_coefficient_context": column_letter(country_headers, "gini_coefficient"),
        "ppp_private_consumption_context": column_letter(country_headers, "ppp_private_consumption"),
        "tax_rate_context": column_letter(country_headers, "tax_rate_assumption"),
        "household_debt_proxy_context": column_letter(country_headers, "household_debt_proxy"),
    }

    formula_columns = {
        "di_ppp_raw",
        "household_debt_effective_raw",
        "gdp_per_capita_ppp_context",
        "gdp_growth_context",
        "gini_coefficient_context",
        "ppp_private_consumption_context",
        "tax_rate_context",
        "household_debt_proxy_context",
    }
    input_columns = {
        "gross_income",
        "rent",
        "utilities",
        "transit_cost",
        "internet_cost",
        "food_cost",
        "housing_burden_raw",
        "household_debt_burden_raw",
        "working_time_pressure_raw",
        "suicide_mental_strain_raw",
        "personal_safety_raw",
        "transit_access_commute_raw",
        "clean_air_raw",
        "water_sanitation_utility_raw",
        "digital_infrastructure_raw",
        "healthcare_quality_raw",
        "education_quality_raw",
        "equal_opportunity_raw",
        "hospitality_belonging_raw",
        "tolerance_pluralism_raw",
        "cultural_public_life_raw",
        "entrepreneurial_dynamism_raw",
        "innovation_research_intensity_raw",
        "investment_signal_raw",
        "administrative_investment_friction_raw",
        "visitor_flow_context_raw",
    }
    context_metadata_columns = {
        "pressure_source_url",
        "pressure_source_tier",
        "pressure_source_date",
        "pressure_notes",
        "viability_source_url",
        "viability_source_tier",
        "viability_source_date",
        "viability_notes",
        "capability_source_url",
        "capability_source_tier",
        "capability_source_date",
        "capability_notes",
        "community_source_url",
        "community_source_tier",
        "community_source_date",
        "community_notes",
        "creative_source_url",
        "creative_source_tier",
        "creative_source_date",
        "creative_notes",
    }

    for row_index, city in enumerate(city_universe, start=2):
        row_values: list[object] = []
        for header in headers:
            if header == "city_id":
                value = city["city_id"]
            elif header == "display_name":
                value = city["display_name"]
            elif header == "country":
                value = city["country"]
            elif header == "cohort":
                value = city["cohort"]
            elif header == "city_type":
                value = city["city_type"]
            elif header == "manifest_status":
                value = city["manifest_status"]
            elif header in country_column_letters:
                value = city_input_formula(
                    row_index,
                    country_column_letters[header],
                    country_count + 1,
                    city_country_letter,
                )
            elif header == "di_ppp_raw":
                value = (
                    f'=IF(OR(${city_gross_income_letter}{row_index}="",${city_rent_letter}{row_index}="",'
                    f'${city_utilities_letter}{row_index}="",${city_transit_cost_letter}{row_index}="",'
                    f'${city_internet_cost_letter}{row_index}="",${city_food_cost_letter}{row_index}="",'
                    f'${city_ppp_context_letter}{row_index}="",${city_tax_context_letter}{row_index}=""),"",'
                    f'(((${city_gross_income_letter}{row_index}*(1-${city_tax_context_letter}{row_index}))-'
                    f'${city_rent_letter}{row_index}-${city_utilities_letter}{row_index}-${city_transit_cost_letter}{row_index}-'
                    f'${city_internet_cost_letter}{row_index}-${city_food_cost_letter}{row_index})/'
                    f'${city_ppp_context_letter}{row_index}))'
                )
            elif header == "household_debt_effective_raw":
                value = (
                    f'=IF(${city_household_debt_burden_letter}{row_index}<>"",'
                    f'${city_household_debt_burden_letter}{row_index},${city_debt_context_letter}{row_index})'
                )
            else:
                value = ""
            row_values.append(value)
        sheet.append(row_values)

    style_input_sheet(sheet, headers, formula_columns, input_columns, context_metadata_columns)
    freeze_and_filter(sheet)
    fit_columns(sheet, max_width=30)


def style_input_sheet(
    sheet: Worksheet,
    headers: list[str],
    formula_columns: set[str],
    input_columns: set[str],
    context_columns: set[str],
) -> None:
    for col_index, header in enumerate(headers, start=1):
        if header in formula_columns:
            fill = FORMULA_FILL
        elif header in input_columns:
            fill = INPUT_FILL
        elif header in context_columns:
            fill = CONTEXT_FILL
        else:
            fill = SUBHEADER_FILL
        for row_index in range(2, sheet.max_row + 1):
            sheet.cell(row=row_index, column=col_index).fill = fill
            sheet.cell(row=row_index, column=col_index).alignment = WRAP


def build_norm_stats_sheet(
    sheet: Worksheet,
    headers: list[str],
    score_inputs: list[dict[str, object]],
    city_input_headers: list[str],
    city_count: int,
) -> dict[str, int]:
    sheet.append(headers)
    style_header_row(sheet, 1)
    source_col_letter = column_letter(headers, "source_column")
    p05_letter = column_letter(headers, "p05")
    p95_letter = column_letter(headers, "p95")
    row_map: dict[str, int] = {}

    for row_index, score_input in enumerate(score_inputs, start=2):
        row_map[str(score_input["input_key"])] = row_index
        source_column = str(score_input["source_column"])
        city_col_letter = column_letter(city_input_headers, source_column)
        range_ref = f"City_Inputs!${city_col_letter}$2:${city_col_letter}${city_count + 1}"
        values = [
            score_input["input_key"],
            score_input["label"],
            source_column,
            score_input["directionality"],
            score_input["coverage_weight"],
            f'=IF(COUNT({range_ref})>=2,PERCENTILE.INC({range_ref},0.05),"")',
            f'=IF(COUNT({range_ref})>=2,PERCENTILE.INC({range_ref},0.95),"")',
            score_input["notes"],
        ]
        sheet.append(values)

    for row_index in range(2, sheet.max_row + 1):
        sheet[f"{p05_letter}{row_index}"].fill = FORMULA_FILL
        sheet[f"{p95_letter}{row_index}"].fill = FORMULA_FILL
    freeze_and_filter(sheet)
    fit_columns(sheet, max_width=36)
    return row_map


def normalized_expression(
    city_input_headers: list[str],
    score_input_rows: dict[str, int],
    norm_headers: list[str],
    input_key: str,
    row_number: int,
) -> str:
    city_col_letter = column_letter(city_input_headers, input_key)
    norm_row = score_input_rows[input_key]
    p05_letter = column_letter(norm_headers, "p05")
    p95_letter = column_letter(norm_headers, "p95")
    value_ref = f"City_Inputs!${city_col_letter}{row_number}"
    p05_ref = f"Norm_Stats!${p05_letter}${norm_row}"
    p95_ref = f"Norm_Stats!${p95_letter}${norm_row}"

    directions = {item["input_key"]: item["directionality"] for item in MANIFEST["scoreInputs"]}
    direction = directions[input_key]
    if direction == "positive":
        return (
            f'IF(OR({value_ref}="",{p05_ref}="",{p95_ref}="",{p95_ref}={p05_ref}),"",'
            f'MAX(0,MIN(100,100*(MIN(MAX({value_ref},{p05_ref}),{p95_ref})-{p05_ref})/({p95_ref}-{p05_ref}))))'
        )
    return (
        f'IF(OR({value_ref}="",{p05_ref}="",{p95_ref}="",{p95_ref}={p05_ref}),"",'
        f'MAX(0,MIN(100,100*({p95_ref}-MIN(MAX({value_ref},{p05_ref}),{p95_ref}))/({p95_ref}-{p05_ref}))))'
    )


def binary_coverage_formula(score_ref: str) -> str:
    return f'=IF({score_ref}="","",1)'


def composite_score_formula(components: list[tuple[str, float]]) -> str:
    numerator = "+".join(f'IF(({expr})="",0,({expr})*{weight})' for expr, weight in components)
    denominator = "+".join(f'IF(({expr})="",0,{weight})' for expr, weight in components)
    return f'=IF(({denominator})=0,"",({numerator})/({denominator}))'


def composite_coverage_formula(components: list[tuple[str, float]]) -> str:
    availability = "+".join(f'IF(({expr})="",0,{weight})' for expr, weight in components)
    return f'=IF(({availability})=0,"",({availability}))'


def weighted_pillar_score_formula(components: list[tuple[str, int]]) -> str:
    numerator = "+".join(f'IF({ref}="",0,{ref}*{weight})' for ref, weight in components)
    denominator = "+".join(f'IF({ref}="",0,{weight})' for ref, weight in components)
    return f'=IF(({denominator})=0,"",({numerator})/({denominator}))'


def weighted_pillar_coverage_formula(components: list[tuple[str, int]], total_weight: int) -> str:
    numerator = "+".join(f'IF({ref}="",0,{ref}*{weight})' for ref, weight in components)
    availability = "+".join(f'IF({ref}="",0,{weight})' for ref, weight in components)
    return f'=IF(({availability})=0,"",({numerator})/{total_weight})'


def build_scores_sheet(
    sheet: Worksheet,
    headers: list[str],
    city_universe: list[dict[str, str]],
    city_input_headers: list[str],
    norm_headers: list[str],
    score_input_rows: dict[str, int],
    thresholds: dict[str, float],
) -> None:
    sheet.append(headers)
    style_header_row(sheet, 1)

    score_header_to_letter = {header: get_column_letter(index + 1) for index, header in enumerate(headers)}
    metric_specs = {
        "pressure_disposable_income_ppp": {"input_key": "di_ppp_raw", "weight": 9, "type": "direct"},
        "pressure_housing_burden": {"input_key": "housing_burden_raw", "weight": 5, "type": "direct"},
        "pressure_household_debt_burden": {"input_key": "household_debt_effective_raw", "weight": 4, "type": "direct"},
        "pressure_working_time_pressure": {"input_key": "working_time_pressure_raw", "weight": 4, "type": "direct"},
        "pressure_suicide_mental_strain": {"input_key": "suicide_mental_strain_raw", "weight": 3, "type": "direct"},
        "viability_personal_safety": {"input_key": "personal_safety_raw", "weight": 5, "type": "direct"},
        "viability_transit_access_commute": {"input_key": "transit_access_commute_raw", "weight": 5, "type": "direct"},
        "viability_clean_air": {"input_key": "clean_air_raw", "weight": 4, "type": "direct"},
        "viability_water_sanitation_utility": {"input_key": "water_sanitation_utility_raw", "weight": 4, "type": "direct"},
        "viability_digital_infrastructure": {"input_key": "digital_infrastructure_raw", "weight": 4, "type": "direct"},
        "capability_healthcare_quality": {"input_key": "healthcare_quality_raw", "weight": 8, "type": "direct"},
        "capability_education_quality": {"input_key": "education_quality_raw", "weight": 6, "type": "direct"},
        "capability_equal_opportunity_distributional_fairness": {
            "weight": 4,
            "type": "composite",
            "components": [
                ("equal_opportunity_raw", 0.7),
                ("gini_coefficient_context", 0.3),
            ],
        },
        "community_hospitality_belonging": {"input_key": "hospitality_belonging_raw", "weight": 5, "type": "direct"},
        "community_tolerance_pluralism": {"input_key": "tolerance_pluralism_raw", "weight": 5, "type": "direct"},
        "community_cultural_historic_public_life_vitality": {
            "input_key": "cultural_public_life_raw",
            "weight": 5,
            "type": "direct",
        },
        "creative_entrepreneurial_dynamism": {"input_key": "entrepreneurial_dynamism_raw", "weight": 6, "type": "direct"},
        "creative_innovation_research_intensity": {
            "input_key": "innovation_research_intensity_raw",
            "weight": 5,
            "type": "direct",
        },
        "creative_economic_vitality_productive_context": {
            "weight": 5,
            "type": "composite",
            "components": [
                ("investment_signal_raw", 0.5),
                ("gdp_per_capita_ppp_context", 0.3),
                ("gdp_growth_context", 0.2),
            ],
        },
        "creative_administrative_investment_friction": {
            "input_key": "administrative_investment_friction_raw",
            "weight": 4,
            "type": "direct",
        },
    }
    pillar_specs = {
        "pressure": [
            ("pressure_disposable_income_ppp", 9),
            ("pressure_housing_burden", 5),
            ("pressure_household_debt_burden", 4),
            ("pressure_working_time_pressure", 4),
            ("pressure_suicide_mental_strain", 3),
        ],
        "viability": [
            ("viability_personal_safety", 5),
            ("viability_transit_access_commute", 5),
            ("viability_clean_air", 4),
            ("viability_water_sanitation_utility", 4),
            ("viability_digital_infrastructure", 4),
        ],
        "capability": [
            ("capability_healthcare_quality", 8),
            ("capability_education_quality", 6),
            ("capability_equal_opportunity_distributional_fairness", 4),
        ],
        "community": [
            ("community_hospitality_belonging", 5),
            ("community_tolerance_pluralism", 5),
            ("community_cultural_historic_public_life_vitality", 5),
        ],
        "creative": [
            ("creative_entrepreneurial_dynamism", 6),
            ("creative_innovation_research_intensity", 5),
            ("creative_economic_vitality_productive_context", 5),
            ("creative_administrative_investment_friction", 4),
        ],
    }

    for row_index, city in enumerate(city_universe, start=2):
        row_values = [""] * len(headers)
        row_values[0] = city["city_id"]
        row_values[1] = city["display_name"]
        row_values[2] = city["country"]
        row_values[3] = city["cohort"]
        row_values[4] = city["manifest_status"]
        row_values[5] = city["city_type"]
        sheet.append(row_values)

        for metric_key, spec in metric_specs.items():
            score_header = f"{metric_key}_score"
            coverage_header = f"{metric_key}_coverage"
            score_cell = f"{score_header_to_letter[score_header]}{row_index}"
            coverage_cell = f"{score_header_to_letter[coverage_header]}{row_index}"

            if spec["type"] == "direct":
                input_key = str(spec["input_key"])
                score_formula = f'={normalized_expression(city_input_headers, score_input_rows, norm_headers, input_key, row_index)}'
                coverage_formula = binary_coverage_formula(score_cell)
            else:
                component_refs = []
                for input_key, weight in spec["components"]:
                    component_formula = normalized_expression(
                        city_input_headers,
                        score_input_rows,
                        norm_headers,
                        input_key,
                        row_index,
                    )
                    component_refs.append((component_formula, weight))
                score_formula = composite_score_formula(component_refs)
                coverage_formula = composite_coverage_formula(component_refs)

            sheet[score_cell] = score_formula
            sheet[coverage_cell] = coverage_formula
            sheet[score_cell].fill = FORMULA_FILL
            sheet[coverage_cell].fill = FORMULA_FILL

        for pillar_name, metrics in pillar_specs.items():
            score_refs = [(f"{score_header_to_letter[f'{metric}_score']}{row_index}", weight) for metric, weight in metrics]
            coverage_refs = [(f"{score_header_to_letter[f'{metric}_coverage']}{row_index}", weight) for metric, weight in metrics]
            pillar_score_header = f"{pillar_name}_pillar_score"
            pillar_coverage_header = f"{pillar_name}_pillar_coverage"
            sheet[f"{score_header_to_letter[pillar_score_header]}{row_index}"] = weighted_pillar_score_formula(score_refs)
            sheet[f"{score_header_to_letter[pillar_coverage_header]}{row_index}"] = weighted_pillar_coverage_formula(
                coverage_refs,
                sum(weight for _, weight in metrics),
            )
            sheet[f"{score_header_to_letter[pillar_score_header]}{row_index}"].fill = FORMULA_FILL
            sheet[f"{score_header_to_letter[pillar_coverage_header]}{row_index}"].fill = FORMULA_FILL

        overall_coverage_cell = f"{score_header_to_letter['overall_weighted_coverage']}{row_index}"
        coverage_grade_cell = f"{score_header_to_letter['coverage_grade']}{row_index}"
        ranking_status_cell = f"{score_header_to_letter['ranking_status']}{row_index}"
        slic_score_cell = f"{score_header_to_letter['slic_score']}{row_index}"
        rank_cell = f"{score_header_to_letter['rank']}{row_index}"

        pressure_coverage = f"{score_header_to_letter['pressure_pillar_coverage']}{row_index}"
        viability_coverage = f"{score_header_to_letter['viability_pillar_coverage']}{row_index}"
        capability_coverage = f"{score_header_to_letter['capability_pillar_coverage']}{row_index}"
        community_coverage = f"{score_header_to_letter['community_pillar_coverage']}{row_index}"
        creative_coverage = f"{score_header_to_letter['creative_pillar_coverage']}{row_index}"

        pressure_score = f"{score_header_to_letter['pressure_pillar_score']}{row_index}"
        viability_score = f"{score_header_to_letter['viability_pillar_score']}{row_index}"
        capability_score = f"{score_header_to_letter['capability_pillar_score']}{row_index}"
        community_score = f"{score_header_to_letter['community_pillar_score']}{row_index}"
        creative_score = f"{score_header_to_letter['creative_pillar_score']}{row_index}"

        sheet[overall_coverage_cell] = (
            f'=IF(AND({pressure_coverage}="",{viability_coverage}="",{capability_coverage}="",'
            f'{community_coverage}="",{creative_coverage}=""),"",'
            f'(({pressure_coverage}*25)+({viability_coverage}*22)+({capability_coverage}*18)+'
            f'({community_coverage}*15)+({creative_coverage}*20))/100)'
        )
        sheet[coverage_grade_cell] = (
            f'=IF({overall_coverage_cell}="","",IF({overall_coverage_cell}>={thresholds["grade_a_min"]},"A",'
            f'IF({overall_coverage_cell}>={thresholds["grade_b_min"]},"B",'
            f'IF({overall_coverage_cell}>={thresholds["grade_c_min"]},"C","Watchlist"))))'
        )
        sheet[ranking_status_cell] = (
            f'=IF(OR({overall_coverage_cell}="",{overall_coverage_cell}<{thresholds["ranked_min_overall"]},'
            f'{pressure_coverage}="",{pressure_coverage}<{thresholds["ranked_min_pillar"]},'
            f'{viability_coverage}="",{viability_coverage}<{thresholds["ranked_min_pillar"]},'
            f'{capability_coverage}="",{capability_coverage}<{thresholds["ranked_min_pillar"]},'
            f'{community_coverage}="",{community_coverage}<{thresholds["ranked_min_pillar"]},'
            f'{creative_coverage}="",{creative_coverage}<{thresholds["ranked_min_pillar"]}),"Watchlist","Ranked")'
        )
        sheet[slic_score_cell] = (
            f'=IF(AND({pressure_score}="",{viability_score}="",{capability_score}="",{community_score}="",{creative_score}=""),"",'
            f'('
            f'IF({pressure_score}="",0,{pressure_score}*25)+'
            f'IF({viability_score}="",0,{viability_score}*22)+'
            f'IF({capability_score}="",0,{capability_score}*18)+'
            f'IF({community_score}="",0,{community_score}*15)+'
            f'IF({creative_score}="",0,{creative_score}*20)'
            f')/('
            f'IF({pressure_score}="",0,25)+'
            f'IF({viability_score}="",0,22)+'
            f'IF({capability_score}="",0,18)+'
            f'IF({community_score}="",0,15)+'
            f'IF({creative_score}="",0,20)'
            f'))'
        )
        sheet[rank_cell] = (
            f'=IF({ranking_status_cell}="Ranked",1+COUNTIFS(${score_header_to_letter["ranking_status"]}$2:${score_header_to_letter["ranking_status"]}${len(city_universe)+1},"Ranked",'
            f'${score_header_to_letter["slic_score"]}$2:${score_header_to_letter["slic_score"]}${len(city_universe)+1},">"&{slic_score_cell}),"")'
        )
        visitor_flow_letter = column_letter(city_input_headers, "visitor_flow_context_raw")
        sheet[f"{score_header_to_letter['visitor_flow_context_raw']}{row_index}"] = f'=City_Inputs!${visitor_flow_letter}{row_index}'

    for row_index in range(2, sheet.max_row + 1):
        for header in headers[6:]:
            sheet[f"{score_header_to_letter[header]}{row_index}"].fill = FORMULA_FILL
            sheet[f"{score_header_to_letter[header]}{row_index}"].alignment = WRAP

    freeze_and_filter(sheet)
    fit_columns(sheet, max_width=28)


def build_leaderboard_sheet(sheet: Worksheet, headers: list[str], score_headers: list[str], city_count: int) -> None:
    title_row = 1
    top_headers_row = 2
    top_start_row = 3
    sheet["A1"] = "Top 100 Ranked Cities"
    sheet["A1"].font = TITLE_FONT
    sheet["A1"].fill = TITLE_FILL
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    for column_index, header in enumerate(headers, start=1):
        sheet.cell(row=top_headers_row, column=column_index, value=header)
    style_header_row(sheet, top_headers_row)

    score_header_to_letter = {header: get_column_letter(index + 1) for index, header in enumerate(score_headers)}

    def leaderboard_formula(target_rank_ref: str, score_field: str) -> str:
        score_col = score_header_to_letter[score_field]
        rank_col = score_header_to_letter["rank"]
        return (
            f'=IFERROR(INDEX(Scores!${score_col}$2:${score_col}${city_count + 1},'
            f'MATCH({target_rank_ref},Scores!${rank_col}$2:${rank_col}${city_count + 1},0)),"")'
        )

    for row_offset in range(100):
        row_number = top_start_row + row_offset
        target_rank_ref = f"$B{row_number}"
        sheet[f"A{row_number}"] = "Top 100"
        sheet[f"B{row_number}"] = row_offset + 1
        sheet[f"C{row_number}"] = leaderboard_formula(target_rank_ref, "rank")
        sheet[f"D{row_number}"] = leaderboard_formula(target_rank_ref, "display_name")
        sheet[f"E{row_number}"] = leaderboard_formula(target_rank_ref, "country")
        sheet[f"F{row_number}"] = leaderboard_formula(target_rank_ref, "cohort")
        sheet[f"G{row_number}"] = leaderboard_formula(target_rank_ref, "manifest_status")
        sheet[f"H{row_number}"] = leaderboard_formula(target_rank_ref, "coverage_grade")
        sheet[f"I{row_number}"] = leaderboard_formula(target_rank_ref, "slic_score")
        sheet[f"J{row_number}"] = leaderboard_formula(target_rank_ref, "pressure_pillar_score")
        sheet[f"K{row_number}"] = leaderboard_formula(target_rank_ref, "viability_pillar_score")
        sheet[f"L{row_number}"] = leaderboard_formula(target_rank_ref, "capability_pillar_score")
        sheet[f"M{row_number}"] = leaderboard_formula(target_rank_ref, "community_pillar_score")
        sheet[f"N{row_number}"] = leaderboard_formula(target_rank_ref, "creative_pillar_score")

    second_title_row = top_start_row + 102
    second_header_row = second_title_row + 1
    second_start_row = second_header_row + 1
    sheet[f"A{second_title_row}"] = "Top 50 Ranked Cities"
    sheet[f"A{second_title_row}"].font = TITLE_FONT
    sheet[f"A{second_title_row}"].fill = TITLE_FILL
    sheet.merge_cells(start_row=second_title_row, start_column=1, end_row=second_title_row, end_column=len(headers))

    for column_index, header in enumerate(headers, start=1):
        sheet.cell(row=second_header_row, column=column_index, value=header)
    style_header_row(sheet, second_header_row)

    for row_offset in range(50):
        row_number = second_start_row + row_offset
        target_rank_ref = f"$B{row_number}"
        sheet[f"A{row_number}"] = "Top 50"
        sheet[f"B{row_number}"] = row_offset + 1
        sheet[f"C{row_number}"] = leaderboard_formula(target_rank_ref, "rank")
        sheet[f"D{row_number}"] = leaderboard_formula(target_rank_ref, "display_name")
        sheet[f"E{row_number}"] = leaderboard_formula(target_rank_ref, "country")
        sheet[f"F{row_number}"] = leaderboard_formula(target_rank_ref, "cohort")
        sheet[f"G{row_number}"] = leaderboard_formula(target_rank_ref, "manifest_status")
        sheet[f"H{row_number}"] = leaderboard_formula(target_rank_ref, "coverage_grade")
        sheet[f"I{row_number}"] = leaderboard_formula(target_rank_ref, "slic_score")
        sheet[f"J{row_number}"] = leaderboard_formula(target_rank_ref, "pressure_pillar_score")
        sheet[f"K{row_number}"] = leaderboard_formula(target_rank_ref, "viability_pillar_score")
        sheet[f"L{row_number}"] = leaderboard_formula(target_rank_ref, "capability_pillar_score")
        sheet[f"M{row_number}"] = leaderboard_formula(target_rank_ref, "community_pillar_score")
        sheet[f"N{row_number}"] = leaderboard_formula(target_rank_ref, "creative_pillar_score")

    freeze_and_filter(sheet, freeze_at="A3")
    fit_columns(sheet, max_width=24)


def build_region_comparison_sheet(sheet: Worksheet, cohorts: list[str], score_headers: list[str], city_count: int) -> None:
    headers = [
        "cohort",
        "ranked_city_count",
        "average_slic_score",
        "average_overall_coverage",
        "average_pressure_score",
        "average_creative_score",
        "top_ranked_city",
    ]
    sheet.append(headers)
    style_header_row(sheet, 1)
    score_header_to_letter = {header: get_column_letter(index + 1) for index, header in enumerate(score_headers)}

    for row_index, cohort in enumerate(cohorts, start=2):
        cohort_ref = f"$A{row_index}"
        cohort_col = score_header_to_letter["cohort"]
        status_col = score_header_to_letter["ranking_status"]
        rank_col = score_header_to_letter["rank"]
        score_col = score_header_to_letter["slic_score"]
        coverage_col = score_header_to_letter["overall_weighted_coverage"]
        pressure_col = score_header_to_letter["pressure_pillar_score"]
        creative_col = score_header_to_letter["creative_pillar_score"]
        display_col = score_header_to_letter["display_name"]

        sheet[f"A{row_index}"] = cohort
        sheet[f"B{row_index}"] = (
            f'=COUNTIFS(Scores!${cohort_col}$2:${cohort_col}${city_count + 1},{cohort_ref},'
            f'Scores!${status_col}$2:${status_col}${city_count + 1},"Ranked")'
        )
        sheet[f"C{row_index}"] = (
            f'=IFERROR(AVERAGEIFS(Scores!${score_col}$2:${score_col}${city_count + 1},'
            f'Scores!${cohort_col}$2:${cohort_col}${city_count + 1},{cohort_ref},'
            f'Scores!${status_col}$2:${status_col}${city_count + 1},"Ranked"),"")'
        )
        sheet[f"D{row_index}"] = (
            f'=IFERROR(AVERAGEIFS(Scores!${coverage_col}$2:${coverage_col}${city_count + 1},'
            f'Scores!${cohort_col}$2:${cohort_col}${city_count + 1},{cohort_ref}),"")'
        )
        sheet[f"E{row_index}"] = (
            f'=IFERROR(AVERAGEIFS(Scores!${pressure_col}$2:${pressure_col}${city_count + 1},'
            f'Scores!${cohort_col}$2:${cohort_col}${city_count + 1},{cohort_ref},'
            f'Scores!${status_col}$2:${status_col}${city_count + 1},"Ranked"),"")'
        )
        sheet[f"F{row_index}"] = (
            f'=IFERROR(AVERAGEIFS(Scores!${creative_col}$2:${creative_col}${city_count + 1},'
            f'Scores!${cohort_col}$2:${cohort_col}${city_count + 1},{cohort_ref},'
            f'Scores!${status_col}$2:${status_col}${city_count + 1},"Ranked"),"")'
        )
        sheet[f"G{row_index}"] = (
            f'=IFERROR(INDEX(Scores!${display_col}$2:${display_col}${city_count + 1},'
            f'MATCH(MINIFS(Scores!${rank_col}$2:${rank_col}${city_count + 1},'
            f'Scores!${cohort_col}$2:${cohort_col}${city_count + 1},{cohort_ref},'
            f'Scores!${status_col}$2:${status_col}${city_count + 1},"Ranked"),'
            f'Scores!${rank_col}$2:${rank_col}${city_count + 1},0)),"")'
        )

    freeze_and_filter(sheet)
    fit_columns(sheet, max_width=28)


def build_thai_scaffold_sheet(sheet: Worksheet) -> None:
    headers = [
        "city_id",
        "display_name",
        "province",
        "province_gdp_per_capita",
        "pollution_index",
        "hospital_count",
        "school_count",
        "infrastructure_index",
        "median_income",
        "longevity",
        "competitiveness_index",
        "thai_slic_score",
        "rank",
        "notes",
    ]
    sheet.append(headers)
    style_header_row(sheet, 1)

    for city_id, display_name, province in THAI_SCAFFOLD_ROWS:
        sheet.append([city_id, display_name, province, "", "", "", "", "", "", "", "", "", "", ""])

    start_row = 2
    end_row = len(THAI_SCAFFOLD_ROWS) + 1
    for row_index in range(start_row, end_row + 1):
        gdp_ref = f"D{row_index}"
        pollution_ref = f"E{row_index}"
        hospital_ref = f"F{row_index}"
        school_ref = f"G{row_index}"
        infrastructure_ref = f"H{row_index}"
        income_ref = f"I{row_index}"
        longevity_ref = f"J{row_index}"
        competitiveness_ref = f"K{row_index}"

        def percentile(ref: str, rng: str, reverse: bool = False) -> str:
            base = f'IF(COUNT({rng})>=2,PERCENTRANK.INC({rng},{ref})*100,"")'
            if not reverse:
                return base
            return f'IF(COUNT({rng})>=2,(1-PERCENTRANK.INC({rng},{ref}))*100,"")'

        gdp_score = percentile(gdp_ref, f"$D$2:$D${end_row}")
        pollution_score = percentile(pollution_ref, f"$E$2:$E${end_row}", reverse=True)
        hospital_score = percentile(hospital_ref, f"$F$2:$F${end_row}")
        school_score = percentile(school_ref, f"$G$2:$G${end_row}")
        infrastructure_score = percentile(infrastructure_ref, f"$H$2:$H${end_row}")
        income_score = percentile(income_ref, f"$I$2:$I${end_row}")
        longevity_score = percentile(longevity_ref, f"$J$2:$J${end_row}")
        competitiveness_score = percentile(competitiveness_ref, f"$K$2:$K${end_row}")

        sheet[f"L{row_index}"] = (
            f'=IF(COUNT(D{row_index}:K{row_index})=0,"",('
            f'IF({gdp_ref}="",0,{gdp_score}*0.15)+'
            f'IF({pollution_ref}="",0,{pollution_score}*0.15)+'
            f'IF({hospital_ref}="",0,{hospital_score}*0.15)+'
            f'IF({school_ref}="",0,{school_score}*0.10)+'
            f'IF({infrastructure_ref}="",0,{infrastructure_score}*0.10)+'
            f'IF({income_ref}="",0,{income_score}*0.15)+'
            f'IF({longevity_ref}="",0,{longevity_score}*0.10)+'
            f'IF({competitiveness_ref}="",0,{competitiveness_score}*0.10)'
            f')/('
            f'IF({gdp_ref}="",0,0.15)+'
            f'IF({pollution_ref}="",0,0.15)+'
            f'IF({hospital_ref}="",0,0.15)+'
            f'IF({school_ref}="",0,0.10)+'
            f'IF({infrastructure_ref}="",0,0.10)+'
            f'IF({income_ref}="",0,0.15)+'
            f'IF({longevity_ref}="",0,0.10)+'
            f'IF({competitiveness_ref}="",0,0.10)))'
        )
        sheet[f"M{row_index}"] = f'=IF(L{row_index}="", "", 1+COUNTIF($L$2:$L${end_row}, ">"&L{row_index}))'
        sheet[f"L{row_index}"].fill = FORMULA_FILL
        sheet[f"M{row_index}"].fill = FORMULA_FILL

    sheet["N2"] = "Thailand-only companion ranking. This sheet is intentionally separate from global SLIC so province and city data can be handled with Thai-specific logic."
    sheet["N2"].alignment = WRAP
    freeze_and_filter(sheet)
    fit_columns(sheet, max_width=28)


def export_city_csv(city_universe: list[dict[str, str]]) -> None:
    headers = ["city_id", "cohort", "manifest_status", "city_type", "country", "inclusion_rationale"]
    for path in (CITY_CSV_PATH, APP_CITY_CSV_PATH):
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=headers)
            writer.writeheader()
            for city in city_universe:
                writer.writerow({header: city[header] for header in headers})


def apply_number_formats(workbook: Workbook, manifest: dict) -> None:
    country_context_sheet = workbook["Country_Context"]
    country_headers = manifest["sheetColumns"]["Country_Context"]
    country_map = header_map(country_headers)
    for row_index in range(2, country_context_sheet.max_row + 1):
        country_context_sheet.cell(row=row_index, column=country_map["tax_rate_assumption"]).number_format = "0.0%"
        country_context_sheet.cell(row=row_index, column=country_map["gdp_growth"]).number_format = "0.0%"
        country_context_sheet.cell(row=row_index, column=country_map["update_date"]).number_format = "yyyy-mm-dd"

    scores_sheet = workbook["Scores"]
    score_headers = manifest["sheetColumns"]["Scores"]
    score_map = header_map(score_headers)
    percent_fields = [
        "pressure_disposable_income_ppp_coverage",
        "pressure_housing_burden_coverage",
        "pressure_household_debt_burden_coverage",
        "pressure_working_time_pressure_coverage",
        "pressure_suicide_mental_strain_coverage",
        "pressure_pillar_coverage",
        "viability_personal_safety_coverage",
        "viability_transit_access_commute_coverage",
        "viability_clean_air_coverage",
        "viability_water_sanitation_utility_coverage",
        "viability_digital_infrastructure_coverage",
        "viability_pillar_coverage",
        "capability_healthcare_quality_coverage",
        "capability_education_quality_coverage",
        "capability_equal_opportunity_distributional_fairness_coverage",
        "capability_pillar_coverage",
        "community_hospitality_belonging_coverage",
        "community_tolerance_pluralism_coverage",
        "community_cultural_historic_public_life_vitality_coverage",
        "community_pillar_coverage",
        "creative_entrepreneurial_dynamism_coverage",
        "creative_innovation_research_intensity_coverage",
        "creative_economic_vitality_productive_context_coverage",
        "creative_administrative_investment_friction_coverage",
        "creative_pillar_coverage",
        "overall_weighted_coverage",
    ]
    for row_index in range(2, scores_sheet.max_row + 1):
        for field in percent_fields:
            scores_sheet.cell(row=row_index, column=score_map[field]).number_format = "0%"
        for field in ("slic_score", "pressure_pillar_score", "viability_pillar_score", "capability_pillar_score", "community_pillar_score", "creative_pillar_score"):
            scores_sheet.cell(row=row_index, column=score_map[field]).number_format = "0.0"


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    manifest = MANIFEST
    city_universe = build_city_universe()
    countries = unique_countries(city_universe)

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    readme = workbook.create_sheet("README")
    write_readme(readme, manifest)

    metric_catalog = workbook.create_sheet("Metric_Catalog")
    write_table(metric_catalog, manifest["sheetColumns"]["Metric_Catalog"], manifest["metricDefinitions"])

    city_universe_sheet = workbook.create_sheet("City_Universe")
    write_table(city_universe_sheet, manifest["sheetColumns"]["City_Universe"], city_universe)

    country_context_rows = [
        {
            "country": country,
            "gdp_per_capita_ppp": "",
            "gdp_growth": "",
            "gini_coefficient": "",
            "ppp_private_consumption": "",
            "tax_rate_assumption": "",
            "household_debt_proxy": "",
            "source_url": "",
            "source_tier": "Tier 3",
            "update_date": "",
            "notes": "Populate official country context inputs before ranking.",
        }
        for country in countries
    ]
    country_context_sheet = workbook.create_sheet("Country_Context")
    write_table(country_context_sheet, manifest["sheetColumns"]["Country_Context"], country_context_rows)

    city_inputs = workbook.create_sheet("City_Inputs")
    build_city_inputs_sheet(
        city_inputs,
        manifest["sheetColumns"]["City_Inputs"],
        city_universe,
        manifest["sheetColumns"]["Country_Context"],
        len(countries),
    )

    norm_stats = workbook.create_sheet("Norm_Stats")
    score_input_rows = build_norm_stats_sheet(
        norm_stats,
        manifest["sheetColumns"]["Norm_Stats"],
        manifest["scoreInputs"],
        manifest["sheetColumns"]["City_Inputs"],
        len(city_universe),
    )

    scores = workbook.create_sheet("Scores")
    build_scores_sheet(
        scores,
        manifest["sheetColumns"]["Scores"],
        city_universe,
        manifest["sheetColumns"]["City_Inputs"],
        manifest["sheetColumns"]["Norm_Stats"],
        score_input_rows,
        manifest["coverageThresholds"],
    )

    leaderboard = workbook.create_sheet("Leaderboard")
    build_leaderboard_sheet(
        leaderboard,
        manifest["sheetColumns"]["Leaderboard"],
        manifest["sheetColumns"]["Scores"],
        len(city_universe),
    )

    region_comparison = workbook.create_sheet(REGION_COMPARISON_SHEET)
    build_region_comparison_sheet(
        region_comparison,
        list(dict.fromkeys(city["cohort"] for city in city_universe)),
        manifest["sheetColumns"]["Scores"],
        len(city_universe),
    )

    thai_scaffold = workbook.create_sheet(THAI_SCAFFOLD_SHEET)
    build_thai_scaffold_sheet(thai_scaffold)

    apply_number_formats(workbook, manifest)
    workbook.save(WORKBOOK_PATH)
    export_city_csv(city_universe)


if __name__ == "__main__":
    main()
